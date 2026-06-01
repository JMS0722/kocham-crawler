from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


EXCEL_COLUMNS = [
    ("No.", "_idx", 6),
    ("회사명", "company_name", 35),
    ("업종", "category", 30),
    ("전화번호", "phone", 18),
    ("이메일", "email", 28),
    ("주소", "address", 45),
    ("제2사무소", "address_2nd", 40),
    ("웹사이트", "website", 30),
    ("직원 수", "employees", 14),
    ("담당자", "members", 35),
    ("회사 소개", "profile", 50),
]


def export_eurocham_excel(data: list[dict], save_path: str, include_profile: bool = True) -> str:
    """EuroCham 기업 데이터를 엑셀 파일로 저장한다."""
    columns = [c for c in EXCEL_COLUMNS if include_profile or c[1] != "profile"]

    wb = Workbook()
    ws = wb.active
    ws.title = "EuroCham Directory"

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, (header, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, item in enumerate(data, 2):
        for col_idx, (_, key, _) in enumerate(columns, 1):
            if key == "_idx":
                value = row_idx - 1
            else:
                value = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(save_path)
    return save_path


def generate_eurocham_filename(prefix: str = "EuroCham_Directory") -> str:
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{now}.xlsx"

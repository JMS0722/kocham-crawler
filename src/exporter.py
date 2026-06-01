from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


EXCEL_COLUMNS = [
    ("No.", "wr_id", 8),
    ("회사명", "company_name", 40),
    ("대표자", "director", 20),
    ("업종", "business_type", 30),
    ("전화번호", "tel", 18),
    ("이메일", "email", 30),
    ("지역", "area", 14),
    ("주소", "address", 50),
    ("홈페이지", "homepage", 30),
    ("직원 수", "staff", 25),
    ("사업 내용", "service", 50),
]


def export_to_excel(data: list[dict], save_path: str) -> str:
    """기업 데이터를 엑셀 파일로 저장한다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KOCHAM 디렉토리"

    # header style
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_font = Font(name="맑은 고딕", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)

    # write headers
    for col_idx, (header, _, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # write data
    for row_idx, item in enumerate(data, 2):
        for col_idx, (_, key, _) in enumerate(EXCEL_COLUMNS, 1):
            value = item.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # freeze header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(save_path)
    return save_path


def generate_filename(prefix: str = "KOCHAM_Directory") -> str:
    """날짜 기반 파일명을 생성한다."""
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{now}.xlsx"
